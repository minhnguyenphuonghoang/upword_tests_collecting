#!/usr/local/bin/python
# -*- coding: utf-8 -*-
"""
@Author: Minh.nguyen
@Created Date: Sunday August 2, 2015
"""
def get_test_data_by_sheet_name(path_to_data_file, sheet_name):
    """Get all data from spread sheet file (only *.xlsx)

    Required: first row will be taken as test cases id (unique attribue) 
    and make it as the primary key to find test data for that test case
    
    Ex: 
    1. Spread sheet format:
    | Test case id (optional title) | header1 (mandatory title) | header2 (mandatory title) |
    | tc_001 | data01_test01 | data02_test02 |
    | tc_002 | data01_test02 | data02_test02 |

    2. Return value
    {
        'tc_001' : {'header1' : 'data01_test01', 'header2' : 'data01_test02'},
        'tc_002' : {'header2' : 'data01_test02', 'header2' : 'data02_test02'}
    }

    """
    
    from openpyxl import load_workbook
    from openpyxl import Workbook
    
    wb = None
    ws = None
    
    try:
        wb = load_workbook(path_to_data_file)
        ws = wb.get_sheet_by_name(sheet_name)
    except:
        assert False, "Could not find excel file located in path: '%s'" % (path_to_data_file)

    if ws is None:
        assert False, "Sheet name '%s' does not exist in excel file " % (sheet_name)

    # get highest row and column
    row_count = ws.max_row + 1
    column_count = ws.max_column + 1
   
    test_data = {}

    for row_index in range(2, row_count):
        current_row_data = {}

        for column_index in range(2, column_count):
            
            # get header and make it as a key to find data
            header = ws.cell(row = 1, column = column_index).value
            # if header is none, skip this row
            if header is None:
                continue
            else:
                header = str(header).lower()
            
            # get data at column index
            data = ws.cell(row = row_index, column = column_index).value or ""

            
            current_row_data.update({header:data})
        # get test case id and mark it as the primary key to find a row data
        test_id = ws.cell(row = row_index, column = 1).value
        # add row data to test data
        test_data.update({test_id:current_row_data})

    return test_data

def write_test_data_by_sheet_name(path_to_data_file, sheet_name, column_name, row_index, data):
    """Get all data from spread sheet file (only *.xlsx)

    Required: first row will be taken as test cases id (unique attribue) 
    and make it as the primary key to find test data for that test case
    
    Ex: 
    1. Spread sheet format:
    | Test case id (optional title) | header1 (mandatory title) | header2 (mandatory title) |
    | tc_001 | data01_test01 | data02_test02 |
    | tc_002 | data01_test02 | data02_test02 |

    2. Return value
    {
        'tc_001' : {'header1' : 'data01_test01', 'header2' : 'data01_test02'},
        'tc_002' : {'header2' : 'data01_test02', 'header2' : 'data02_test02'}
    }

    """
    
    from openpyxl import load_workbook
    from openpyxl import Workbook
    
    wb = None
    ws = None
    
    try:
        wb = load_workbook(path_to_data_file)
    except:
        assert False, "Could not find excel file located in path: '%s'" % (path_to_data_file)
    try:
        ws = wb.get_sheet_by_name(sheet_name)
    except:
        ws = wb.create_sheet(0)
        ws.title = sheet_name

    ws[column_name+str(row_index)] = data

    wb.save(path_to_data_file)

def is_question_exist(path_to_data_file, sheet_name, question):
    """

    """
    
    from openpyxl import load_workbook
    from openpyxl import Workbook
    
    wb = None
    ws = None
    
    try:
        wb = load_workbook(path_to_data_file)
    except:
        assert False, "Could not find excel file located in path: '%s'" % (path_to_data_file)
    try:
        ws = wb.get_sheet_by_name(sheet_name)
    except:
        return False
        
    if ws is None:
        assert False, "Sheet name '%s' does not exist in excel file " % (sheet_name)

    # get highest row and column
    row_count = ws.max_row + 1
    
    index = 1

    for row_index in range(1, row_count):
        current_data = ws.cell(row = row_index, column = 1).value

        if current_data is None: continue
        if type(current_data) == str or type(current_data) == unicode:
            current_data = current_data.encode('utf-8')
        else:
            current_data = str(current_data)

        if (current_data == question):
            return True

    return False


def write_test_data_by_sheet_name_in_latest_row(path_to_data_file, sheet_name, column_name, data):
    """

    """
    
    from openpyxl import load_workbook
    from openpyxl import Workbook
    
    wb = None
    ws = None
    
    try:
        wb = load_workbook(path_to_data_file)
    except:
        assert False, "Could not find excel file located in path: '%s'" % (path_to_data_file)
    try:
        ws = wb.get_sheet_by_name(sheet_name)
    except:
        ws = wb.create_sheet(0)
        ws.title = sheet_name

    row_count = ws.max_row + 1

    ws[column_name+str(row_count)] = data

    wb.save(path_to_data_file)

    return row_count


# if __name__ == '__main__':
#     a = is_question_exist("/Users/minhshome/upword_tests_collecting/text.xlsx","test","111aaa")
#     print a







